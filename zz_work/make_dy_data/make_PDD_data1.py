import os
import pandas as pd
import numpy as np
import re
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils.dataframe import dataframe_to_rows

# 获取程序所在目录
current_dir = os.path.dirname(os.path.abspath(__file__))

# 读取file文件夹下的所有csv文件
def read_sales_files():
    file_dir = os.path.join(current_dir, 'file')
    
    all_files = glob.glob(os.path.join(file_dir, '*.csv'))
    
    if not all_files:
        print("file文件夹中没有找到csv文件")
        return None
    
    # 存储所有读取成功的数据
    all_data = []
    
    for file_path in all_files:
        file_name = os.path.basename(file_path)
        try:
            # 尝试不同编码读取文件
            try:
                df = pd.read_csv(file_path, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(file_path, encoding='gb2312')
                except UnicodeDecodeError:
                    df = pd.read_csv(file_path, encoding='gbk')
            
            # 检查必要的列是否存在
            required_columns = ["订单号", "订单状态", "商品数量(件)", "商品id", 
                               "商家编码-规格维度", "售后状态", "商家备注", "商家实收金额(元)"]
            
            if not all(col in df.columns for col in required_columns):
                print(f"文件 {file_name} 缺少必要的列，跳过处理")
                continue
            
            # 只保留需要的列
            df = df[required_columns].copy()
            
            # 添加文件名列
            df['文件名'] = file_name
            
            # 处理商家实收金额，去除千位分隔符
            df['商家实收金额(元)'] = df['商家实收金额(元)'].astype(str).str.replace(',', '').astype(float)
            
            # 清理商品id，只保留数字
            df['商品id'] = df['商品id'].astype(str).apply(lambda x: re.sub(r'\D', '', x))
            
            # 拆分商家编码-规格维度
            df['商家编码名字'] = df['商家编码-规格维度'].astype(str).apply(
                lambda x: x.split('-')[0] if '-' in x and pd.notna(x) and x != 'nan' else '')
            
            df['商家编码数量'] = df['商家编码-规格维度'].astype(str).apply(
                lambda x: x.split('-')[1] if '-' in x and len(x.split('-')) > 1 and pd.notna(x) and x != 'nan' else '')
            
            all_data.append(df)
            
        except Exception as e:
            print(f"读取文件 {file_name} 时出错: {str(e)}")
    
    if not all_data:
        print("没有成功读取任何销售数据文件")
        return None
    
    # 合并所有数据
    sales_data = pd.concat(all_data, ignore_index=True)
    
    # 根据订单号去重
    sales_data = sales_data.drop_duplicates(subset=['订单号'])
    
    # 筛选订单状态和售后状态
    valid_order_status = ["待发货", "已发货，待收货", "已收货"]
    valid_after_sale_status = ["无售后或售后取消", "售后处理中"]
    
    sales_data = sales_data[
        sales_data['订单状态'].isin(valid_order_status) & 
        sales_data['售后状态'].isin(valid_after_sale_status)
    ]
    
    return sales_data

# 读取km文件夹下的csv文件
def read_km_files():
    km_dir = os.path.join(current_dir, 'km')
    all_files = glob.glob(os.path.join(km_dir, '*.csv'))
    
    all_data = []
    
    for file_path in all_files:
        file_name = os.path.basename(file_path)
        try:
            # 尝试不同编码读取文件
            try:
                df = pd.read_csv(file_path, encoding='utf-8', skiprows=6)
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(file_path, encoding='gb2312', skiprows=6)
                except UnicodeDecodeError:
                    df = pd.read_csv(file_path, encoding='gbk', skiprows=6)
            
            # 检查必要的列是否存在
            if "主商家编码" not in df.columns or "商品名称" not in df.columns:
                print(f"快卖文件 {file_name} 缺少必要的列，跳过处理")
                continue
                
            # 只保留需要的列
            df = df[["主商家编码", "商品名称"]].copy()
            
            all_data.append(df)
        except Exception as e:
            print(f"读取快卖文件 {file_name} 时出错: {str(e)}")
    
    if not all_data:
        print("没有成功读取任何快卖数据文件")
        return None
    
    # 合并所有数据
    km_data = pd.concat(all_data, ignore_index=True)
    
    # 删除空值行
    km_data = km_data.dropna(subset=["主商家编码", "商品名称"])
    
    # 根据主商家编码去重
    km_data = km_data.drop_duplicates(subset=['主商家编码'])
    
    return km_data

# 读取产品数据表格
def read_product_data():
    product_file = os.path.join(current_dir, '产品数据表格.xlsx')
    
    try:
        # 获取所有表格名称
        xls = pd.ExcelFile(product_file)
        sheet_names = [name for name in xls.sheet_names if name[-1] == '组']
        
        all_data = []
        
        for sheet_name in sheet_names:
            df = pd.read_excel(product_file, sheet_name=sheet_name, usecols=["店铺", "姓名", "产品ID"])
            df['组名字'] = sheet_name
            all_data.append(df)
        
        if not all_data:
            print("产品数据表格中没有找到以'组'结尾的表格")
            return None
        
        # 合并所有数据
        product_data = pd.concat(all_data, ignore_index=True)
        
        # 清理产品ID，只保留数字
        product_data['产品ID'] = product_data['产品ID'].astype(str).apply(lambda x: re.sub(r'\D', '', x))
        
        # 根据产品ID去重
        product_data = product_data.drop_duplicates(subset=['产品ID'])
        
        return product_data
        
    except Exception as e:
        print(f"读取产品数据表格时出错: {str(e)}")
        return None

# 读取dtb文件夹下的文件
def read_dtb_files():
    dtb_dir = os.path.join(current_dir, 'dtb')
    
    # 检查文件夹是否存在
    if not os.path.exists(dtb_dir):
        print("dtb文件夹不存在，跳过处理")
        return None
    
    all_files = glob.glob(os.path.join(dtb_dir, '*.xlsx'))
    
    if not all_files:
        print("dtb文件夹中没有xlsx文件，跳过处理")
        return None
    
    all_data = []
    
    for file_path in all_files:
        file_name = os.path.basename(file_path)
        try:
            df = pd.read_excel(file_path, usecols=["订单号"])
            # 确保订单号是字符串类型
            df['订单号'] = df['订单号'].astype(str)
            # 清理订单号格式，保留原始格式以便匹配
            all_data.append(df)
        except Exception as e:
            print(f"读取dtb文件 {file_name} 时出错: {str(e)}")
    
    if not all_data:
        return None
    
    # 合并所有数据
    dtb_data = pd.concat(all_data, ignore_index=True)
    return dtb_data

# 主函数
def main():
    # 检查必要的文件夹结构
    print("检查程序运行环境...")
    
    # 检查file文件夹
    file_dir = os.path.join(current_dir, 'file')
    if not os.path.exists(file_dir):
        print(f"警告：file文件夹不存在，正在创建...")
        try:
            os.makedirs(file_dir)
            print(f"已创建file文件夹: {file_dir}")
        except Exception as e:
            print(f"创建file文件夹失败: {str(e)}")
    
    # 检查km文件夹
    km_dir = os.path.join(current_dir, 'km')
    if not os.path.exists(km_dir):
        print(f"警告：km文件夹不存在，正在创建...")
        try:
            os.makedirs(km_dir)
            print(f"已创建km文件夹: {km_dir}")
        except Exception as e:
            print(f"创建km文件夹失败: {str(e)}")
    
    # 检查dtb文件夹
    dtb_dir = os.path.join(current_dir, 'dtb')
    if not os.path.exists(dtb_dir):
        print(f"警告：dtb文件夹不存在，正在创建...")
        try:
            os.makedirs(dtb_dir)
            print(f"已创建dtb文件夹: {dtb_dir}")
        except Exception as e:
            print(f"创建dtb文件夹失败: {str(e)}")
    
    # 检查产品数据表格
    product_file = os.path.join(current_dir, '产品数据表格.xlsx')
    if not os.path.exists(product_file):
        print(f"警告：产品数据表格.xlsx不存在，请确保文件放在正确位置")
        print(f"预期路径: {product_file}")
    
    # 读取销售数据
    print("正在读取销售数据...")
    sales_data = read_sales_files()
    if sales_data is None:
        print("无法继续处理，请确保:")
        print("1. file文件夹中有CSV格式的销售数据文件")
        print("2. CSV文件包含必要的列（订单号、商品id等）")
        print("按任意键退出程序...")
        input()
        return
    
    # 读取快卖数据
    print("正在读取快卖数据...")
    km_data = read_km_files()
    
    # 读取产品数据
    print("正在读取产品数据...")
    product_data = read_product_data()
    
    # 读取dtb数据
    print("正在读取dtb数据...")
    dtb_data = read_dtb_files()
    
    # 合并数据
    if km_data is not None:
        # 左连接销售数据和快卖数据
        sales_data = pd.merge(
            sales_data, 
            km_data, 
            how='left', 
            left_on='商家编码名字', 
            right_on='主商家编码'
        )
    
    if product_data is not None:
        # 左连接销售数据和产品数据
        sales_data = pd.merge(
            sales_data, 
            product_data, 
            how='left', 
            left_on='商品id', 
            right_on='产品ID'
        )
    
    # 处理数据
    # 添加类型列，默认为"真实"
    sales_data['类型'] = '真实'
    
    # 根据商家备注修改类型
    sales_data.loc[sales_data['商家备注'].str.contains(r'[Vv][-_]', na=False), '类型'] = '放单'
    sales_data.loc[sales_data['商家备注'].str.contains(r'[Gg][-_]', na=False), '类型'] = '刷单'
    
    # 根据dtb数据修改类型
    if dtb_data is not None:
        # 确保销量数据中的订单号也是字符串类型
        sales_data['订单号'] = sales_data['订单号'].astype(str)
        
        # 标准化两边的订单号格式
        # 移除所有空格
        sales_data['订单号_标准'] = sales_data['订单号'].str.strip()
        dtb_data['订单号_标准'] = dtb_data['订单号'].str.strip()
        
        dtb_orders = dtb_data['订单号_标准'].unique()
        sales_data.loc[sales_data['订单号_标准'].isin(dtb_orders), '类型'] = '放单'
        
        # 删除临时列
        sales_data.drop('订单号_标准', axis=1, inplace=True)
    
    # 添加快递数量列
    sales_data['快递数量'] = 1
    sales_data.loc[sales_data['类型'] == '刷单', '快递数量'] = 0
    
    # 补全数据
    # 补全姓名
    sales_data['姓名'] = sales_data['姓名'].fillna('没有写名字')
    
    # 补全商品名称
    sales_data['商品名称'] = sales_data['商品名称'].fillna('没有对应到产品')
    
    # 补全店铺和组名字
    for col in ['店铺', '组名字']:
        if col in sales_data.columns:
            # 对每个文件名，找出该列最常见的值
            for file_name in sales_data['文件名'].unique():
                file_data = sales_data[sales_data['文件名'] == file_name]
                most_common = file_data[col].mode()
                if not most_common.empty:
                    most_common_value = most_common[0]
                    # 只填充该文件中的空值
                    mask = (sales_data['文件名'] == file_name) & (sales_data[col].isna())
                    sales_data.loc[mask, col] = most_common_value
    
    # 保存原始数据
    sales_data.to_csv(os.path.join(current_dir, '原始数据表格.csv'), index=False, encoding='utf-8')
    
    print("原始数据已保存到原始数据表格.csv")
    
    # 生成商品ID数据透视表
    print("正在生成商品ID数据表格...")
    
    # 创建商品ID数据透视表
    product_id_pivot = pd.pivot_table(
        sales_data,
        values=['商家实收金额(元)', '快递数量'],
        index=['商品id', '商品名称', '店铺', '姓名', '组名字', '类型'],
        aggfunc=np.sum,
        fill_value=0
    )
    
    # 重置索引，将索引转换为列
    product_id_pivot = product_id_pivot.reset_index()
    
    # 将金额列四舍五入到2位小数
    product_id_pivot['商家实收金额(元)'] = product_id_pivot['商家实收金额(元)'].round(2)
    
    # 将快递数量转为整数
    product_id_pivot['快递数量'] = product_id_pivot['快递数量'].astype(int)
    
    # 按商品id和类型排序
    product_id_pivot = product_id_pivot.sort_values(by=['商品id', '类型'])
    
    # 保存商品ID数据表格
    product_id_pivot.to_csv(os.path.join(current_dir, '商品ID数据表格.csv'), index=False, encoding='utf-8')
    
    print("商品ID数据表格已保存到商品ID数据表格.csv")
    
    # 创建透视表
    # 第一个表：按姓名和类型汇总金额
    pivot1 = pd.pivot_table(
        sales_data, 
        values='商家实收金额(元)', 
        index=['姓名'], 
        columns=['类型'], 
        aggfunc=np.sum, 
        fill_value=0
    )
    
    # 第二个表：按组名字和类型汇总金额
    pivot2 = pd.pivot_table(
        sales_data, 
        values='商家实收金额(元)', 
        index=['组名字'], 
        columns=['类型'], 
        aggfunc=np.sum, 
        fill_value=0
    )
    
    # 计算每个组的人数
    group_counts = sales_data.groupby('组名字')['姓名'].nunique()
    
    # 添加组平均销售额列
    if '真实' in pivot2.columns:
        pivot2['组平均销售额'] = (pivot2['真实'] / group_counts).round(2)
    
    # 根据"真实"列从高到低排序
    if '真实' in pivot1.columns:
        pivot1 = pivot1.sort_values(by='真实', ascending=False)
    
    if '真实' in pivot2.columns:
        pivot2 = pivot2.sort_values(by='真实', ascending=False)
    
    # 将所有金额列四舍五入到2位小数
    for col in pivot1.columns:
        pivot1[col] = pivot1[col].round(2)
    
    for col in pivot2.columns:
        pivot2[col] = pivot2[col].round(2)
    
    # 分离直通车和非直通车数据
    non_ztc_data = sales_data[~sales_data['组名字'].str.contains('直通车', na=False)]
    ztc_data = sales_data[sales_data['组名字'].str.contains('直通车', na=False)]
    
    # 第三个表：非直通车数据按商品名称汇总（简化版）
    # 先创建基础透视表
    pivot3_base = pd.pivot_table(
        non_ztc_data, 
        values=['商家实收金额(元)', '快递数量'], 
        index=['商品名称'], 
        columns=['类型'], 
        aggfunc=np.sum, 
        fill_value=0
    )
    
    # 创建简化版透视表
    pivot3 = pd.DataFrame(index=pivot3_base.index)
    
    # 添加真实销售额
    if ('商家实收金额(元)', '真实') in pivot3_base.columns:
        pivot3['销售额'] = pivot3_base[('商家实收金额(元)', '真实')].round(2)
    else:
        pivot3['销售额'] = 0
    
    # 添加放单销售额
    if ('商家实收金额(元)', '放单') in pivot3_base.columns:
        pivot3['放单'] = pivot3_base[('商家实收金额(元)', '放单')].round(2)
    else:
        pivot3['放单'] = 0
    
    # 计算总快递数量（所有类型）
    pivot3['快递数量'] = non_ztc_data.groupby('商品名称')['快递数量'].sum().astype(int)
    
    # 计算客单价
    pivot3['客单价'] = (pivot3['销售额'] / pivot3['快递数量'].replace(0, np.nan)).round(2)
    
    # 按销售额降序排序
    pivot3 = pivot3.sort_values(by='销售额', ascending=False)
    
    # 第四个表：直通车数据按商品名称汇总（简化版）
    if not ztc_data.empty:
        # 先创建基础透视表
        pivot4_base = pd.pivot_table(
            ztc_data, 
            values=['商家实收金额(元)', '快递数量'], 
            index=['商品名称'], 
            columns=['类型'], 
            aggfunc=np.sum, 
            fill_value=0
        )
        
        # 创建简化版透视表
        pivot4 = pd.DataFrame(index=pivot4_base.index)
        
        # 添加真实销售额
        if ('商家实收金额(元)', '真实') in pivot4_base.columns:
            pivot4['销售额'] = pivot4_base[('商家实收金额(元)', '真实')].round(2)
        else:
            pivot4['销售额'] = 0
        
        # 添加放单销售额
        if ('商家实收金额(元)', '放单') in pivot4_base.columns:
            pivot4['放单'] = pivot4_base[('商家实收金额(元)', '放单')].round(2)
        else:
            pivot4['放单'] = 0
        
        # 计算总快递数量（所有类型）
        pivot4['快递数量'] = ztc_data.groupby('商品名称')['快递数量'].sum().astype(int)
        
        # 计算客单价
        pivot4['客单价'] = (pivot4['销售额'] / pivot4['快递数量'].replace(0, np.nan)).round(2)
        
        # 按销售额降序排序
        pivot4 = pivot4.sort_values(by='销售额', ascending=False)
    else:
        pivot4 = pd.DataFrame()
    
    # 创建Excel工作簿并保存结果
    wb = Workbook()
    
    # 使用默认的Sheet
    ws = wb.active
    ws.title = "汇总数据"
    
    # 当前列的位置
    current_col = 1
    
    # 添加第一个表：按姓名汇总
    # 添加标题
    ws.cell(row=1, column=current_col).value = "按姓名汇总"
    ws.cell(row=1, column=current_col).font = Font(bold=True, size=14)
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(pivot1, index=True, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=2+r_idx, column=current_col+c_idx).value = value
    
    # 更新当前列位置
    current_col += len(pivot1.columns) + 2  # 数据列数 + 索引列 + 空列
    
    # 添加第二个表：按组名字汇总
    # 添加标题
    ws.cell(row=1, column=current_col).value = "按组名字汇总"
    ws.cell(row=1, column=current_col).font = Font(bold=True, size=14)
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(pivot2, index=True, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=2+r_idx, column=current_col+c_idx).value = value
    
    # 更新当前列位置
    current_col += len(pivot2.columns) + 2  # 数据列数 + 索引列 + 空列
    
    # 添加第三个表：非直通车商品汇总
    # 添加标题
    ws.cell(row=1, column=current_col).value = "非直通车商品汇总"
    ws.cell(row=1, column=current_col).font = Font(bold=True, size=14)
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(pivot3, index=True, header=True)):
        for c_idx, value in enumerate(row):
            ws.cell(row=2+r_idx, column=current_col+c_idx).value = value
    
    # 更新当前列位置
    current_col += len(pivot3.columns) + 2  # 数据列数 + 索引列 + 空列
    
    # 添加第四个表：直通车商品汇总
    if not pivot4.empty:
        # 添加标题
        ws.cell(row=1, column=current_col).value = "直通车商品汇总"
        ws.cell(row=1, column=current_col).font = Font(bold=True, size=14)
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(pivot4, index=True, header=True)):
            for c_idx, value in enumerate(row):
                ws.cell(row=2+r_idx, column=current_col+c_idx).value = value
    
    # 设置列宽
    for column in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=column)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = adjusted_width
    
    # 添加边框和设置数字格式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_row = 2
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:  # 只为有值的单元格添加边框
                cell.border = thin_border
                # 对数值类型的单元格设置数字格式
                if isinstance(cell.value, (int, float)) and row > 2 and col > 1:
                    # 获取当前列的标题
                    header_cell = ws.cell(row=header_row, column=col)
                    header_text = str(header_cell.value) if header_cell.value is not None else ""
                    
                    # 如果列标题是"快递数量"，则使用整数格式
                    if "快递数量" in header_text:
                        cell.number_format = '0'
                    else:
                        cell.number_format = '0.00'
    
    # 美化表格 - 最后应用颜色样式
    # 定义颜色
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 蓝色标题
    alt_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # 浅蓝色交替行
    index_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # 中蓝色索引列
    section_title_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")  # 区域标题背景色
    
    # 设置区域标题格式
    for col in [1, len(pivot1.columns) + 2, len(pivot1.columns) + 2 + len(pivot2.columns) + 2, 
                len(pivot1.columns) + 2 + len(pivot2.columns) + 2 + len(pivot3.columns) + 2]:
        if col <= ws.max_column:
            cell = ws.cell(row=1, column=col)
            if cell.value:  # 确保单元格有值
                cell.font = Font(bold=True, size=14)
                cell.fill = section_title_fill
    
    # 设置每个表格的样式
    # 处理第一个表格：按姓名汇总
    start_col = 1
    
    # 设置标题行格式
    for col in range(start_col, start_col + len(pivot1.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
    
    # 设置数据行格式
    for row in range(header_row + 1, header_row + 1 + len(pivot1)):
        # 设置索引列格式
        index_cell = ws.cell(row=row, column=start_col)
        if index_cell.value is not None:
            index_cell.font = Font(bold=True)
            index_cell.fill = index_fill
        
        # 设置交替行颜色
        if (row - header_row - 1) % 2 == 0:  # 偶数行
            for col in range(start_col + 1, start_col + len(pivot1.columns) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_row_fill
    
    # 处理第二个表格：按组名字汇总
    start_col = len(pivot1.columns) + 2
    
    # 设置标题行格式
    for col in range(start_col, start_col + len(pivot2.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
    
    # 设置数据行格式
    for row in range(header_row + 1, header_row + 1 + len(pivot2)):
        # 设置索引列格式
        index_cell = ws.cell(row=row, column=start_col)
        if index_cell.value is not None:
            index_cell.font = Font(bold=True)
            index_cell.fill = index_fill
        
        # 设置交替行颜色
        if (row - header_row - 1) % 2 == 0:  # 偶数行
            for col in range(start_col + 1, start_col + len(pivot2.columns) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_row_fill
    
    # 处理第三个表格：非直通车商品汇总
    start_col = len(pivot1.columns) + 2 + len(pivot2.columns) + 2
    
    # 设置标题行格式
    for col in range(start_col, start_col + len(pivot3.columns) + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
    
    # 设置数据行格式
    for row in range(header_row + 1, header_row + 1 + len(pivot3)):
        # 设置索引列格式
        index_cell = ws.cell(row=row, column=start_col)
        if index_cell.value is not None:
            index_cell.font = Font(bold=True)
            index_cell.fill = index_fill
        
        # 设置交替行颜色
        if (row - header_row - 1) % 2 == 0:  # 偶数行
            for col in range(start_col + 1, start_col + len(pivot3.columns) + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.fill = alt_row_fill
    
    # 处理第四个表格：直通车商品汇总
    if not pivot4.empty:
        start_col = len(pivot1.columns) + 2 + len(pivot2.columns) + 2 + len(pivot3.columns) + 2
        
        # 设置标题行格式
        for col in range(start_col, start_col + len(pivot4.columns) + 1):
            cell = ws.cell(row=header_row, column=col)
            if cell.value is not None:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
        
        # 设置数据行格式
        for row in range(header_row + 1, header_row + 1 + len(pivot4)):
            # 设置索引列格式
            index_cell = ws.cell(row=row, column=start_col)
            if index_cell.value is not None:
                index_cell.font = Font(bold=True)
                index_cell.fill = index_fill
            
            # 设置交替行颜色
            if (row - header_row - 1) % 2 == 0:  # 偶数行
                for col in range(start_col + 1, start_col + len(pivot4.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.fill = alt_row_fill
    
    # 保存Excel文件
    excel_file = os.path.join(current_dir, '数据汇总.xlsx')
    wb.save(excel_file)
    
    print(f"数据汇总已保存到 {excel_file}")
    print("处理完成！按任意键退出...")
    input()

# 程序入口
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        print("按任意键退出...")
        input()